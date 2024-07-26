from openpyxl import load_workbook


class Excel_func:

    def __init__(self, filename, sheet_number):
        self.file = filename
        self.sheet = sheet_number

    def row_count(self):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        return sheet.max_row

    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    def read_data(self, row_number, column_number):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    def write_data(self, row_number, column_number, data):
        work_book = load_workbook(self.file)
        sheet = work_book[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        work_book.save(self.file)